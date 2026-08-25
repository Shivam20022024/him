import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelService:
    @classmethod
    def get_candidate_file_path(cls, org_id: str):
        return f"candidates_{org_id}.xlsx"
        
    @classmethod
    def get_call_results_file_path(cls, org_id: str):
        return f"candidate_call_results_{org_id}.xlsx"

    HEADERS = ["Candidate ID", "Name", "Email", "Score", "Status", "Interest", "Timestamp"]
    
    CALL_RESULTS_HEADERS = [
        "Candidate ID", "Candidate Name", "Email", "Phone", "Job Role", 
        "Total Experience", "Relevant Experience", "Employment Status", 
        "Joining Availability", "Interview Availability", "Interest Status", 
        "Transcript", "Recording URL", "Created Timestamp"
    ]

    @classmethod
    def save_call_result_to_excel(cls, data: dict, org_id: str):
        """
        Appends real call result data to candidate_call_results_{org_id}.xlsx.
        """
        try:
            if not org_id:
                logger.warning("No org_id provided for save_call_result_to_excel")
                return False
                
            file_path = cls.get_call_results_file_path(org_id)
            
            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = "Call Results"
                ws.append(cls.CALL_RESULTS_HEADERS)
                wb.save(file_path)

            wb = load_workbook(file_path)
            ws = wb.active

            # Extract fields with safe defaults
            # Mapping from candidate model/dict to Excel columns
            row = [
                data.get("id") or data.get("candidate_id") or "N/A",
                data.get("name") or "N/A",
                data.get("email") or "N/A",
                data.get("phone") or "N/A",
                data.get("role") or data.get("job_role") or "N/A",
                data.get("total_experience") or data.get("experience_years") or "N/A",
                data.get("relevant_experience") or "N/A",
                data.get("employment_status") or "N/A",
                data.get("joining_availability") or data.get("availability") or "N/A",
                data.get("interview_availability") or "N/A",
                data.get("interest") or data.get("interested") or data.get("interest_status") or "N/A",
                data.get("transcript") or "N/A",
                data.get("recording_url") or "N/A",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]

            duplicate_found = False
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == row[0]: # Column 1 is ID
                    for c in range(len(row)):
                        ws.cell(row=r, column=c+1).value = row[c]
                    duplicate_found = True
                    break
            
            if not duplicate_found:
                ws.append(row)
                
            wb.save(file_path)
            logger.info(f"Successfully saved/updated call result to Excel for {data.get('name')}")
            return True
        except Exception as e:
            logger.error(f"Failed to save call result to Excel: {str(e)}")
            return False

    @classmethod
    def reset_candidate_excel(cls, org_id: str):
        """Recreates the candidate workbook with only the header row."""
        try:
            file_path = cls.get_candidate_file_path(org_id)
            wb = Workbook()
            ws = wb.active
            ws.title = "Candidates"
            ws.append(cls.HEADERS)
            wb.save(file_path)
            logger.info(f"Reset Excel file: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Failed to reset Excel: {str(e)}")
            return False

    @classmethod
    def update_candidate_excel(cls, candidate_data: dict, org_id: str):
        """
        Updates the candidates_{org_id}.xlsx file with new candidate data.
        If the file doesn't exist, it creates one with headers.
        """
        try:
            if not org_id:
                logger.warning("No org_id provided for update_candidate_excel")
                return False
                
            file_path = cls.get_candidate_file_path(org_id)
            
            if not os.path.exists(file_path):
                # Create a new workbook and add headers
                wb = Workbook()
                ws = wb.active
                ws.title = "Candidates"
                ws.append(cls.HEADERS)
                wb.save(file_path)
                logger.info(f"Created new Excel file: {file_path}")

            # Load existing workbook
            wb = load_workbook(file_path)
            ws = wb.active

            # Prepare row data
            # Map candidate_data keys to headers
            row = [
                candidate_data.get("id", "N/A"),
                candidate_data.get("name", "N/A"),
                candidate_data.get("email", "N/A"),
                f"{candidate_data.get('resume_score', 0)}%",
                candidate_data.get("status", "pending"),
                candidate_data.get("interest", "pending"),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]

            # Check for duplicates (based on Candidate ID)
            # For a simple append-only log, skip this if we want full history.
            # But the requirement said 'Ensure no duplicate entries'.
            duplicate_found = False
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=1).value == row[0]: # Column 1 is ID
                    # Update existing row
                    for c in range(len(row)):
                        ws.cell(row=r, column=c+1).value = row[c]
                    duplicate_found = True
                    logger.info(f"Updated existing entry in Excel for: {row[1]}")
                    break
            
            if not duplicate_found:
                ws.append(row)
                logger.info(f"Added new entry to Excel: {row[1]}")

            wb.save(file_path)
            logger.info("Excel updated successfully.")
            return True
        except Exception as e:
            logger.error(f"Failed to update Excel: {str(e)}")
            return False
